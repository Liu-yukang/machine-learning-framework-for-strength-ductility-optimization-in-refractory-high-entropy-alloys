import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import KFold
from sklearn.metrics import r2_score, mean_squared_error
from itertools import combinations
from tqdm import tqdm
import matplotlib.pyplot as plt
import os
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

np.random.seed(42)

selected_features = ['number', 'ρ', 'δB', 'D.B', 'v', 'Smix', 'Hmix', 'Gmix', 'Λ', 'γ', 'ƞ', 'μ', 'A', 'F', 'a', 'Δa']
target = "Ys"

def load_data(file_path):
    print("加载数据...")
    data = pd.read_excel(file_path).loc[:, ~data.columns.duplicated()]

    missing = data[selected_features + [target]].isnull().sum()
    print("\n缺失值：")
    print(missing)
    
    for col in selected_features + [target]:
        if data[col].isnull().any():
            data[col] = data[col].fillna(data[col].median())
    
    return data

def evaluate_model(X, y, n_splits=10, n_repeats=10):
    res = {'r2': [], 'rmse': []}
    
    for _ in tqdm(range(n_repeats), desc="CV", leave=False):
        kf = KFold(n_splits, shuffle=True, random_state=np.random.randint(1, 1000))
        fold_r2, fold_rmse = [], []
        
        for train_idx, test_idx in kf.split(X):
            X_tr, X_te = X.iloc[train_idx], X.iloc[test_idx]
            y_tr, y_te = y.iloc[train_idx], y.iloc[test_idx]

            mean_tr = X_tr.mean()
            std_tr = X_tr.std()
            
            if any(std_tr == 0):
                print(f"警告: 特征 {std_tr[std_tr == 0].index.tolist()} 无波动，跳过标准化")
                X_tr_s, X_te_s = X_tr.copy(), X_te.copy()
            else:
                X_tr_s = (X_tr - mean_tr) / std_tr
                X_te_s = (X_te - mean_tr) / std_tr

            model = RandomForestRegressor(100, max_depth=5, random_state=42, n_jobs=-1)
            model.fit(X_tr_s, y_tr)

            y_pr = model.predict(X_te_s)
            fold_r2.append(r2_score(y_te, y_pr))
            fold_rmse.append(np.sqrt(mean_squared_error(y_te, y_pr)))
        
        res['r2'].append(np.mean(fold_r2))
        res['rmse'].append(np.mean(fold_rmse))
    
    return {
        'mean_r2': np.mean(res['r2']),
        'std_r2': np.std(res['r2']),
        'mean_rmse': np.mean(res['rmse']),
        'std_rmse': np.std(res['rmse'])
    }

def process_comb(features, data, n_repeats=10):
    X = data[list(features)]
    y = data[target]
    res = evaluate_model(X, y, n_repeats=n_repeats)
    return {
        'features': features,
        'mean_r2': res['mean_r2'],
        'std_r2': res['std_r2'],
        'mean_rmse': res['mean_rmse'],
        'std_rmse': res['std_rmse']
    }

def select_best(data, min_f=3, max_f=7, n_repeats=10):
    combs = []
    for k in range(min_f, max_f + 1):
        combs.extend(list(combinations(selected_features, k)))
    
    total = len(combs)
    print(f"\n评估组合数：{total}")
    
    results = []
    for f in tqdm(combs, desc="进度", unit="组", total=total):
        results.append(process_comb(f, data, n_repeats))
    
    return sorted(results, key=lambda x: x['mean_r2'], reverse=True)

def visualize_top(top_res, n=10):
    plt.figure(figsize=(12, 8))
    top_n = top_res[:n]
    labels = [str(c['features']) for c in top_n]
    r2 = [c['mean_r2'] for c in top_n]
    r2_err = [c['std_r2'] for c in top_n]
    
    plt.barh(labels, r2, xerr=r2_err, capsize=5)
    plt.xlabel('Mean R²')
    plt.ylabel('Feature Combinations')
    plt.title('Top Combinations (R² with Std)')
    plt.grid(axis='x', linestyle='--', alpha=0.7)
    plt.tight_layout()
    
    if not os.path.exists('feature_results'):
        os.makedirs('feature_results')
    plt.savefig('feature_results/top_combs.png', dpi=300)
    plt.close()

def save_excel(results, output):
    print(f"\n保存结果至 {output}...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    
    headers = ["Rank", "Num Features", "Features", "Mean R²", "R² Std", "Mean RMSE", "RMSE Std"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    for rank, res in enumerate(results, 1):
        row = [
            rank,
            len(res['features']),
            ", ".join(res['features']),
            res['mean_r2'],
            res['std_r2'],
            res['mean_rmse'],
            res['std_rmse']
        ]
        ws.append(row)
 
    for col in ws.columns:
        max_len = 0
        col_let = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_let].width = max_len + 2
    
    ws.auto_filter.ref = ws.dimensions
    wb.save(output)
    print(f"已保存至 {output}")

def main():
    file_path = r'C:\Users\Liu\Desktop\666666.xlsx'
    data = load_data(file_path)
    
    best_res = select_best(data, min_f=3, max_f=7, n_repeats=10)
    
    print("\nTop 10 combinations:")
    print(f"{'Rank':<5} {'Features':<120} {'Mean R²':<10} {'R² Std':<10} {'Mean RMSE':<10} {'RMSE Std'}")
    for i, res in enumerate(best_res[:10], 1):
        print(f"{i:<5} {str(res['features']):<120} {res['mean_r2']:.4f} {res['std_r2']:.4f} {res['mean_rmse']:.4f} {res['std_rmse']:.4f}")
    
    visualize_top(best_res)
    
    output = r'C:\Users\Liu\Desktop\最佳特征组合.xlsx'
    save_excel(best_res, output)
    
    print("\n完成")

if __name__ == "__main__":
    main()
